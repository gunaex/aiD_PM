import openpyxl
from sqlalchemy.orm import Session
import models

def export_weekly_report(db: Session, project_id: int, template_path: str, output_path: str):
    """
    สร้าง Weekly Report จาก Template โดยเติมข้อมูล PB Curve
    
    Args:
        db: Database session
        project_id: ID ของโปรเจกต์
        template_path: ที่อยู่ไฟล์ Template
        output_path: ที่อยู่ไฟล์ Output
    
    Returns:
        output_path: ที่อยู่ไฟล์ที่สร้างเสร็จ
    """
    # โหลดไฟล์ต้นฉบับที่คุณแชร์มา
    wb = openpyxl.load_workbook(template_path)
    
    # ดึงข้อมูล Snapshots เพื่อไปทำ PB Curve
    snapshots = db.query(models.WeeklySnapshot).filter(
        models.WeeklySnapshot.project_id == project_id
    ).order_by(models.WeeklySnapshot.week_number).all()
    
    if "PB Curve" in wb.sheetnames:
        sheet = wb["PB Curve"]
        # เริ่มหยอดข้อมูลที่แถว 41 ตามพิกัดในไฟล์ของคุณ
        start_row = 41
        for i, snap in enumerate(snapshots):
            sheet.cell(row=start_row + i, column=2).value = snap.week_number  # Column B
            sheet.cell(row=start_row + i, column=3).value = snap.plan_acc    # Column C
            sheet.cell(row=start_row + i, column=4).value = snap.actual_acc  # Column D

    wb.save(output_path)
    return output_path


def export_daily_progress(db: Session, project_id: int, template_path: str, output_path: str):
    """
    สร้าง Daily Progress Report จาก Template
    
    Args:
        db: Database session
        project_id: ID ของโปรเจกต์
        template_path: ที่อยู่ไฟล์ Template
        output_path: ที่อยู่ไฟล์ Output
    
    Returns:
        output_path: ที่อยู่ไฟล์ที่สร้างเสร็จ
    """
    # โหลดไฟล์ต้นฉบับ
    wb = openpyxl.load_workbook(template_path)
    
    # ดึงข้อมูล Tasks ของโปรเจกต์
    tasks = db.query(models.Task).filter(
        models.Task.project_id == project_id
    ).all()
    
    # สามารถเพิ่ม logic การเติมข้อมูลตามโครงสร้างของ Template
    # ตัวอย่าง: เติมข้อมูลใน Sheet แรก
    if wb.sheetnames:
        sheet = wb[wb.sheetnames[0]]
        # เพิ่ม logic การเติมข้อมูลตามต้องการ
        # sheet.cell(row=X, column=Y).value = data
    
    wb.save(output_path)
    return output_path

